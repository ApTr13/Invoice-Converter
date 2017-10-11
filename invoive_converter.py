# INVOICE CONVERTER - Convert Invoice from pdf to excel
# By - Aptr13(https://github.com/ApTr13)

#******IMPORTANT******
#Before running the code install these two libraries:
#$ pip install PyPDF2
#$ pip install openpyxl

# Import libraries 
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader
from openpyxl import Workbook

# Set input and output file names 
input_file = 'invoice.pdf'
output_file = 'invoice.xlsx'

# Open & Read PDF
pdf_file = open(input_file,'rb')
input_pdf = PyPDF2.PdfFileReader(pdf_file)

# Declare Headings of PDF file
main_list = ['From',
 'To',
 'Invoice Number',
 'Order Number',
 'Invoice Date',
 'Due Date',
 'Total Due',
 'Quantity',
 'Service',
 'Rate',
 'Adjust',
 'Sub Total',
 '!"#$%&#']


# New workbook in openpyxl 
wb = Workbook()
ws = wb.active

# Write Headings into excel file
row_num=1
column_num=1
for i in range(len(main_list)-1):
        field = main_list[i]
        ws.cell(row=row_num, column=column_num, value=field)
        column_num += 1

# Count total page number of PDF file
total_pages = input_pdf.getNumPages()

# Extract data from PDF and Write it into excel file
row_num = 2
for i in range(total_pages):
        page = input_pdf.getPage(i)
        page_content = page.extractText()
        column_num = 1
        for i in range(len(main_list)-1):
                field = main_list[i]
                next_field = main_list[i+1]
                # Find position of fields from extracted text of PDF file
                field_pos = page_content.find(field)
                next_field_pos = page_content.find(next_field)
                # Find position of field values from extracted text of PDF file
                field_value_start_pos = field_pos+len(field)
                field_value_end_pos = next_field_pos
                # Extract field values
                field_value = page_content[field_value_start_pos:field_value_end_pos]
                # Write field values into Excel
                ws.cell(row = row_num, column = column_num, value = field_value)
                column_num += 1
        row_num += 1

pdf_file.close()

# Save excel file
wb.save(output_file)

